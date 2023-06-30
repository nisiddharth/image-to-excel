from PIL import Image
from openpyxl import Workbook
from openpyxl.styles import PatternFill

# Open the PNG image file
image = Image.open('Book1.webp')

# Convert the image to RGB mode
image = image.convert('RGB')

# Get the original size of the image
original_width, original_height = image.size

# Ask the user for the desired output dimensions
output_width = int(input("Enter the output width: "))
output_height = int(input("Enter the output height: "))

# Calculate the width and height ratios
width_ratio = original_width / output_width
height_ratio = original_height / output_height

# Create a new Excel workbook
workbook = Workbook()
sheet = workbook.active

# Iterate through each cell in the output dimensions
for y in range(output_height):
    for x in range(output_width):
        # Calculate the corresponding pixel coordinates range
        start_x = int(x * width_ratio)
        end_x = int((x + 1) * width_ratio)
        start_y = int(y * height_ratio)
        end_y = int((y + 1) * height_ratio)

        # Initialize variables for accumulating RGB values
        total_r = 0
        total_g = 0
        total_b = 0

        # Iterate through the pixel coordinates range
        for px in range(start_x, end_x):
            for py in range(start_y, end_y):
                # Get the RGB values of the pixel
                r, g, b = image.getpixel((px, py))

                # Accumulate the RGB values
                total_r += r
                total_g += g
                total_b += b

        # Calculate the average RGB values
        average_r = int(total_r / ((end_x - start_x) * (end_y - start_y)))
        average_g = int(total_g / ((end_x - start_x) * (end_y - start_y)))
        average_b = int(total_b / ((end_x - start_x) * (end_y - start_y)))

        # Convert the RGB values to a hex color code
        hex_color = f'{average_r:02x}{average_g:02x}{average_b:02x}'

        # Create a pattern fill with the hex color
        fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type='solid')

        # Set the fill for the current cell
        cell = sheet.cell(row=y+1, column=x+1)
        cell.fill = fill

# Save the Excel workbook
workbook.save('output1.xlsx')
